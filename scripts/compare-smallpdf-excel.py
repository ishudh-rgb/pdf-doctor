"""Compare PDF Doctor Excel output against SmallPDF reference export."""

import sys
import openpyxl


def load_workbook_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data = {name: list(wb[name].iter_rows(values_only=True)) for name in wb.sheetnames}
    wb.close()
    return data


def row_label(row):
    for cell in row:
        if cell is None or str(cell).strip() == "":
            continue
        text = str(cell).strip()
        if text.replace(",", "").isdigit():
            continue
        if text in ("PARTICULARS", "A", "REVENUE", "-") or "YEAR" in text:
            continue
        return text[:80]
    return ""


def norm(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "").strip()
    if text == "-":
        return "-"
    if text.isdigit():
        return int(text)
    return text


def compare(reference_path, candidate_path):
    reference = load_workbook_rows(reference_path)
    candidate = load_workbook_rows(candidate_path)

    print(f"Reference: {reference_path}")
    print(f"Candidate: {candidate_path}")
    print()

    total_match = 0
    total_mismatch = 0

    for sheet in reference:
        ref_rows = reference[sheet]
        cand_rows = candidate.get(sheet, [])
        print(
            f"{sheet}: rows {len(cand_rows)}/{len(ref_rows)}, "
            f"cols {len(cand_rows[0]) if cand_rows else 0}/{len(ref_rows[0]) if ref_rows else 0}"
        )

        ref_by_label = {row_label(row): row for row in ref_rows if row_label(row)}
        cand_by_label = {row_label(row): row for row in cand_rows if row_label(row)}

        sheet_match = 0
        sheet_mismatch = 0

        for label, ref_row in ref_by_label.items():
            cand_row = cand_by_label.get(label)
            if cand_row is None:
                sheet_mismatch += 1
                print(f"  missing row: {label[:60]}")
                continue

            for col in range(min(len(ref_row), len(cand_row))):
                ref_val = norm(ref_row[col])
                cand_val = norm(cand_row[col])
                if ref_val is None and cand_val is None:
                    continue
                if isinstance(ref_val, (int, float)) or ref_val == "-" or isinstance(
                    cand_val, (int, float)
                ) or cand_val == "-":
                    if ref_val == cand_val:
                        sheet_match += 1
                    elif ref_val is not None or cand_val is not None:
                        sheet_mismatch += 1

        total_match += sheet_match
        total_mismatch += sheet_mismatch
        print(f"  numeric cells: match={sheet_match} mismatch={sheet_mismatch}")
        print()

    print(f"TOTAL numeric match={total_match} mismatch={total_mismatch}")
    return total_mismatch == 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: compare-smallpdf-excel.py <reference.xlsx> <candidate.xlsx>")
        sys.exit(1)

    ok = compare(sys.argv[1], sys.argv[2])
    sys.exit(0 if ok else 1)
